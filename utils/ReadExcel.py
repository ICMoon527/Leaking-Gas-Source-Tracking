import openpyxl
import pandas as pd

class ReadExcel():
    def __init__(self, filepath) -> None:
        # self.data = openpyxl.load_workbook(filepath, read_only=True)['Sheet1']

    # def readSD(self, row, S, D):  # S79D61
    #     row = row + 2
    #     col = (S-1) * 61 + 6 + D
    #     for datas in self.data.iter_rows(min_row=row, min_col=col, max_row=row, max_col=col):  # 按行滚动
    #         col_size = len(datas)
    #         for i in range(col_size):
    #             return datas[i].value

    # def readS(self, row, S):
    #     row = row + 2
    #     row_data_of_S = []
    #     col = (S-1) * 61 + 7
    #     for datas in self.data.iter_rows(min_row=row, max_row=row, min_col=col, max_col=col+60):
    #         for i in range(len(datas)):
    #             row_data_of_S.append(datas[i].value)
    #     return row_data_of_S
        print('reading excel...\n')
        self.read_file = pd.read_excel(filepath)
        self.filepath = filepath

    def cvt2csv(self):
        path = self.filepath.split('.')[0]+'.csv'
        print('Saving to {}\n'.format(path))
        self.read_file.to_csv(path, index=None, header=True)
        print('show: ')
        df = pd.DataFrame(pd.read_csv(path))
        print(df)

if __name__ == '__main__':
    object = ReadExcel('data/Second_part.xlsx')
    object.cvt2csv()